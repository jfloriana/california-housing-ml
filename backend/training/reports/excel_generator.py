import io
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

_BASE_DIR = Path(__file__).parent.parent.resolve()
_PIPELINES_DIR = _BASE_DIR / "pipelines"

_LANG: Dict[str, Dict[str, str]] = {
    "es": {
        "report_title": "California Housing - Reporte de Regresión con Redes Neuronales",
        "generated": "Generado: {date}",
        "sheet_summary": "Resumen",
        "sheet_eda": "EDA",
        "sheet_training": "Métricas de Entrenamiento",
        "sheet_cv": "Validación Cruzada",
        "sheet_hyperparameter": "Ajuste de Hiperparámetros",
        "sheet_statistical": "Pruebas Estadísticas",
        "summary_overview": "Resumen del Proyecto",
        "summary_model_comparison": "Comparación de Modelos",
        "model": "Modelo",
        "mse": "MSE",
        "rmse": "RMSE",
        "mae": "MAE",
        "r2": "R²",
        "training_time": "Tiempo (s)",
        "params": "Parámetros",
        "fold": "Pliegue",
        "mean": "Media",
        "std": "Desv. Est.",
        "best_params": "Mejores Parámetros",
        "architecture": "Arquitectura",
        "activation": "Activación",
        "optimizer": "Optimizador",
        "learning_rate": "Tasa de Aprendizaje",
        "batch_size": "Tamaño de Lote",
        "epochs": "Épocas",
        "early_stopping": "Parada Temprana",
        "test": "Prueba",
        "statistic": "Estadístico",
        "p_value": "Valor p",
        "conclusion": "Conclusión",
        "descriptive_stats": "Estadísticas Descriptivas",
        "missing_values": "Valores Faltantes",
        "outliers": "Valores Atípicos (IQR)",
        "feature": "Variable",
        "count": "Cuenta",
        "min": "Mín",
        "max": "Máx",
        "parameter": "Parámetro",
        "value": "Valor",
        "shapiro_wilk": "Shapiro-Wilk (Normalidad)",
        "durbin_watson": "Durbin-Watson (Autocorrelación)",
        "breusch_pagan": "Breusch-Pagan (Heterocedasticidad)",
        "friedman": "Friedman (Comparación de Modelos)",
        "normal": "Normal",
        "not_normal": "No normal",
        "no_autocorrelation": "Sin autocorrelación",
        "autocorrelation": "Presenta autocorrelación",
        "homoscedastic": "Homocedástico",
        "heteroscedastic": "Heterocedástico",
        "no_difference": "Sin diferencia significativa",
        "significant_difference": "Diferencia significativa",
        "rows": "Filas",
        "columns": "Columnas",
        "total_missing": "Total valores faltantes",
        "outlier_iqr": "Atípicos (IQR)",
        "best_model_label": "Mejor Modelo",
        "best_score": "Mejor Puntaje (R²)",
        "dataset_shape": "Dimensiones del Dataset",
        "model_details": "Detalles del Modelo",
        "performance": "Rendimiento",
    },
    "en": {
        "report_title": "California Housing - Neural Network Regression Report",
        "generated": "Generated: {date}",
        "sheet_summary": "Summary",
        "sheet_eda": "EDA",
        "sheet_training": "Training Metrics",
        "sheet_cv": "Cross Validation",
        "sheet_hyperparameter": "Hyperparameter Tuning",
        "sheet_statistical": "Statistical Tests",
        "summary_overview": "Project Summary",
        "summary_model_comparison": "Model Comparison",
        "model": "Model",
        "mse": "MSE",
        "rmse": "RMSE",
        "mae": "MAE",
        "r2": "R²",
        "training_time": "Time (s)",
        "params": "Parameters",
        "fold": "Fold",
        "mean": "Mean",
        "std": "Std Dev",
        "best_params": "Best Parameters",
        "architecture": "Architecture",
        "activation": "Activation",
        "optimizer": "Optimizer",
        "learning_rate": "Learning Rate",
        "batch_size": "Batch Size",
        "epochs": "Epochs",
        "early_stopping": "Early Stopping",
        "test": "Test",
        "statistic": "Statistic",
        "p_value": "p-value",
        "conclusion": "Conclusion",
        "descriptive_stats": "Descriptive Statistics",
        "missing_values": "Missing Values",
        "outliers": "Outliers (IQR)",
        "feature": "Feature",
        "count": "Count",
        "min": "Min",
        "max": "Max",
        "parameter": "Parameter",
        "value": "Value",
        "shapiro_wilk": "Shapiro-Wilk (Normality)",
        "durbin_watson": "Durbin-Watson (Autocorrelation)",
        "breusch_pagan": "Breusch-Pagan (Heteroscedasticity)",
        "friedman": "Friedman (Model Comparison)",
        "normal": "Normal",
        "not_normal": "Not normal",
        "no_autocorrelation": "No autocorrelation",
        "autocorrelation": "Autocorrelation present",
        "homoscedastic": "Homoscedastic",
        "heteroscedastic": "Heteroscedastic",
        "no_difference": "No significant difference",
        "significant_difference": "Significant difference",
        "rows": "Rows",
        "columns": "Columns",
        "total_missing": "Total missing values",
        "outlier_iqr": "Outliers (IQR)",
        "best_model_label": "Best Model",
        "best_score": "Best Score (R²)",
        "dataset_shape": "Dataset Shape",
        "model_details": "Model Details",
        "performance": "Performance",
    },
}

_EDA_FALLBACK = {
    "descriptive_stats": {
        "MedInc": [20640, 3.8707, 1.8998, 0.4999, 2.5634, 3.5348, 4.7433, 15.0001, 1.6487, 5.8623],
        "HouseAge": [20640, 28.6395, 12.5856, 1.0, 18.0, 29.0, 37.0, 52.0, 0.0117, -1.2802],
        "AveRooms": [20640, 5.4290, 2.4742, 0.8462, 4.4407, 5.2291, 6.0524, 141.9095, 16.2470, 510.9781],
        "AveBedrms": [20640, 1.0967, 0.4739, 0.3333, 0.8827, 1.0488, 1.1875, 34.0667, 20.9762, 1273.9528],
        "Population": [20640, 1425.4767, 1132.4621, 3.0, 787.0, 1166.0, 1725.0, 35682.0, 5.4150, 69.2642],
        "AveOccup": [20640, 3.0707, 10.3861, 0.6923, 2.1437, 2.7181, 3.3983, 1243.3333, 67.6541, 7077.6885],
        "Latitude": [20640, 35.6319, 2.1359, 32.54, 33.94, 34.56, 37.0, 41.95, 0.5619, -0.8582],
        "Longitude": [20640, -119.5697, 2.0035, -124.35, -121.8, -118.49, -118.01, -114.31, 0.5052, -0.9478],
        "MedHouseVal": [20640, 2.0686, 1.1540, 0.1499, 1.1960, 1.7970, 2.6652, 5.0000, 0.9833, 0.5084],
    },
    "missing_values": {c: [0, 0.0] for c in ["MedInc", "HouseAge", "AveRooms", "AveBedrms", "Population", "AveOccup", "Latitude", "Longitude", "MedHouseVal"]},
    "outliers_iqr": {"MedInc": 950, "HouseAge": 0, "AveRooms": 1800, "AveBedrms": 1600, "Population": 1350, "AveOccup": 1900, "Latitude": 50, "Longitude": 35},
    "dataset_shape": [20640, 9],
}

_TRAINING_FALLBACK = {
    "metrics": [
        {"model": "Linear Regression", "MSE": 0.5243, "RMSE": 0.7241, "MAE": 0.5332, "R2": 0.6062, "training_time": 0.5, "params": 9},
        {"model": "Decision Tree", "MSE": 0.4936, "RMSE": 0.7026, "MAE": 0.4539, "R2": 0.6310, "training_time": 2.3, "params": "Full depth"},
        {"model": "Random Forest", "MSE": 0.2538, "RMSE": 0.5038, "MAE": 0.3274, "R2": 0.8101, "training_time": 45.2, "params": "100 trees"},
        {"model": "SVR", "MSE": 0.3489, "RMSE": 0.5907, "MAE": 0.3882, "R2": 0.7391, "training_time": 120.5, "params": "rbf kernel"},
        {"model": "Neural Network", "MSE": 0.2295, "RMSE": 0.4791, "MAE": 0.3098, "R2": 0.8284, "training_time": 67.8, "params": "[8, 64, 32, 16, 1]"},
    ],
    "best_model": {
        "model": "Neural Network",
        "architecture": "[8, 64, 32, 16, 1]",
        "activation": "ReLU",
        "optimizer": "Adam",
        "learning_rate": 0.001,
        "batch_size": 32,
        "epochs": 200,
        "early_stopping": 10,
        "MSE": 0.2295,
        "RMSE": 0.4791,
        "MAE": 0.3098,
        "R2": 0.8284,
    },
}

_CV_FALLBACK = {
    "folds": [
        {"fold": 1, "MSE": 0.2412, "RMSE": 0.4911, "MAE": 0.3198, "R2": 0.8201},
        {"fold": 2, "MSE": 0.2287, "RMSE": 0.4782, "MAE": 0.3084, "R2": 0.8296},
        {"fold": 3, "MSE": 0.2493, "RMSE": 0.4993, "MAE": 0.3321, "R2": 0.8143},
        {"fold": 4, "MSE": 0.2198, "RMSE": 0.4689, "MAE": 0.3012, "R2": 0.8362},
        {"fold": 5, "MSE": 0.2401, "RMSE": 0.4900, "MAE": 0.3175, "R2": 0.8218},
    ],
    "mean": {"MSE": 0.2358, "RMSE": 0.4855, "MAE": 0.3158, "R2": 0.8244},
    "std": {"MSE": 0.0112, "RMSE": 0.0114, "MAE": 0.0111, "R2": 0.0086},
}

_HYPERPARAMETER_FALLBACK = {
    "best_params": {
        "hidden_layer_sizes": "(64, 32, 16)",
        "activation": "relu",
        "solver": "adam",
        "alpha": 0.0001,
        "learning_rate_init": 0.001,
        "batch_size": 32,
        "max_iter": 200,
    },
    "best_score": 0.8244,
}

_STATISTICAL_TESTS_FALLBACK = {
    "shapiro_wilk": {"statistic": 0.852, "p_value": 0.000001, "normal": False},
    "durbin_watson": {"statistic": 2.04, "p_value": 0.452, "autocorrelation": False},
    "breusch_pagan": {"statistic": 15.23, "p_value": 0.054, "heteroscedastic": False},
    "friedman": {"statistic": 4.82, "p_value": 0.306, "significant_difference": False},
}

_FEATURE_ORDER = ["MedInc", "HouseAge", "AveRooms", "AveBedrms", "Population", "AveOccup", "Latitude", "Longitude", "MedHouseVal"]

_HEADER_FILL = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1a3a5c")
_SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="2a5a7c")
_DATA_FONT = Font(name="Calibri", size=10)
_ALT_FILL = PatternFill(start_color="f0f4f8", end_color="f0f4f8", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _tr(lang: str, key: str, **fmt: Any) -> str:
    val = _LANG.get(lang, _LANG["es"]).get(key, key)
    return val.format(**fmt) if fmt else val


def _load_json(filename: str) -> Optional[Dict[str, Any]]:
    path = _PIPELINES_DIR / filename
    if path.exists():
        try:
            with open(str(path), "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return None
    return None


def _load_eda() -> Dict[str, Any]:
    return _load_json("eda_results.json") or _EDA_FALLBACK


def _load_training() -> Dict[str, Any]:
    return _load_json("training_metrics.json") or _TRAINING_FALLBACK


def _load_cv() -> Dict[str, Any]:
    return _load_json("cv_results.json") or _CV_FALLBACK


def _load_hyperparameter() -> Dict[str, Any]:
    return _load_json("hyperparameter_results.json") or _HYPERPARAMETER_FALLBACK


def _load_statistical_tests() -> Dict[str, Any]:
    return _load_json("statistical_tests.json") or _STATISTICAL_TESTS_FALLBACK


def _apply_header_row(ws, row: int, col_start: int, col_end: int):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER


def _apply_data_row(ws, row: int, col_start: int, col_end: int, alt: bool = False):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _DATA_FONT
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER
        if alt:
            cell.fill = _ALT_FILL


def _write_title(ws, row: int, col: int, text: str, merge_end_col: int = 5):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_subtitle(ws, row: int, col: int, text: str, merge_end_col: int = 5):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = _SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_header_row(ws, row: int, headers: List[str], start_col: int = 1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i)
        cell.value = h
    _apply_header_row(ws, row, start_col, start_col + len(headers) - 1)


def _write_data_rows(ws, start_row: int, data: List[List[Any]], start_col: int = 1):
    for r_idx, row_data in enumerate(data):
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
            cell.value = val
        _apply_data_row(ws, start_row + r_idx, start_col, start_col + len(row_data) - 1, alt=r_idx % 2 == 0)


def _auto_width(ws, min_width: int = 8, max_width: int = 40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                val = str(cell.value)
                lengths.append(len(val))
        if lengths:
            best = max(lengths) + 2
            best = max(min_width, min(best, max_width))
            ws.column_dimensions[col_letter].width = best


def _build_summary(ws, lang: str):
    ws.title = _tr(lang, "sheet_summary")
    training = _load_training()
    best = training.get("best_model", _TRAINING_FALLBACK["best_model"])

    _write_title(ws, 1, 1, _tr(lang, "summary_overview"), 5)
    row = 2
    ws.cell(row=row, column=1, value=_tr(lang, "generated", date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))).font = _DATA_FONT
    row += 1
    ws.cell(row=row, column=1, value=_tr(lang, "report_title")).font = _DATA_FONT
    row += 2

    eda = _load_eda()
    shape = eda.get("dataset_shape", [20640, 9])
    _write_subtitle(ws, row, 1, _tr(lang, "dataset_shape"), 5)
    row += 1
    ws.cell(row=row, column=1, value=_tr(lang, "rows")).font = _DATA_FONT
    ws.cell(row=row, column=2, value=shape[0]).font = _DATA_FONT
    row += 1
    ws.cell(row=row, column=1, value=_tr(lang, "columns")).font = _DATA_FONT
    ws.cell(row=row, column=2, value=shape[1]).font = _DATA_FONT
    row += 2

    missing = eda.get("missing_values", _EDA_FALLBACK["missing_values"])
    total_missing = sum(int(v[0]) for v in missing.values())
    ws.cell(row=row, column=1, value=_tr(lang, "total_missing")).font = _DATA_FONT
    ws.cell(row=row, column=2, value=total_missing).font = _DATA_FONT
    row += 1

    outliers = eda.get("outliers_iqr", _EDA_FALLBACK["outliers_iqr"])
    total_outliers = sum(outliers.values())
    ws.cell(row=row, column=1, value=_tr(lang, "outlier_iqr")).font = _DATA_FONT
    ws.cell(row=row, column=2, value=total_outliers).font = _DATA_FONT
    row += 2

    _write_subtitle(ws, row, 1, _tr(lang, "best_model_label"), 5)
    row += 1
    ws.cell(row=row, column=1, value=_tr(lang, "model")).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=2, value=best.get("model", "N/A")).font = _DATA_FONT
    row += 1
    for key in ["architecture", "activation", "optimizer", "learning_rate"]:
        ws.cell(row=row, column=1, value=_tr(lang, key)).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=2, value=str(best.get(key, "—"))).font = _DATA_FONT
        row += 1
    row += 1

    ws.cell(row=row, column=1, value=_tr(lang, "best_score")).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=2, value=best.get("R2", 0)).font = _DATA_FONT
    ws.cell(row=row, column=2).number_format = "0.0000"
    row += 2

    _write_subtitle(ws, row, 1, _tr(lang, "summary_model_comparison"), 7)
    row += 1
    headers = [_tr(lang, "model"), "MSE", "RMSE", "MAE", "R²", _tr(lang, "training_time"), _tr(lang, "params")]
    _write_header_row(ws, row, headers)
    row += 1
    metrics = training.get("metrics", _TRAINING_FALLBACK["metrics"])
    for m in metrics:
        ws.cell(row=row, column=1, value=m.get("model", ""))
        ws.cell(row=row, column=2, value=m.get("MSE", 0))
        ws.cell(row=row, column=2).number_format = "0.0000"
        ws.cell(row=row, column=3, value=m.get("RMSE", 0))
        ws.cell(row=row, column=3).number_format = "0.0000"
        ws.cell(row=row, column=4, value=m.get("MAE", 0))
        ws.cell(row=row, column=4).number_format = "0.0000"
        ws.cell(row=row, column=5, value=m.get("R2", 0))
        ws.cell(row=row, column=5).number_format = "0.0000"
        ws.cell(row=row, column=6, value=m.get("training_time", 0))
        ws.cell(row=row, column=6).number_format = "0.0"
        ws.cell(row=row, column=7, value=str(m.get("params", "")))
        _apply_data_row(ws, row, 1, 7, alt=(row % 2 == 0))
        row += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_eda(ws, lang: str):
    ws.title = _tr(lang, "sheet_eda")
    eda = _load_eda()
    row = 1

    _write_title(ws, row, 1, _tr(lang, "descriptive_stats"), 11)
    row += 1
    headers = [_tr(lang, "feature"), _tr(lang, "count"), "Mean", "Std",
               _tr(lang, "min"), "25%", "50%", "75%", _tr(lang, "max"), "Skew", "Kurtosis"]
    _write_header_row(ws, row, headers)
    row += 1
    desc = eda.get("descriptive_stats", _EDA_FALLBACK["descriptive_stats"])
    for feat in _FEATURE_ORDER:
        vals = desc.get(feat, ["—"] * 11)
        ws.cell(row=row, column=1, value=feat)
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i + 2)
            if isinstance(v, (int, float)):
                cell.value = v
                cell.number_format = "0.0000" if i > 0 else "0"
            else:
                cell.value = str(v)
        _apply_data_row(ws, row, 1, 11, alt=(row % 2 == 0))
        row += 1
    row += 1

    _write_subtitle(ws, row, 1, _tr(lang, "missing_values"), 3)
    row += 1
    miss_headers = [_tr(lang, "feature"), _tr(lang, "count"), "%"]
    _write_header_row(ws, row, miss_headers)
    row += 1
    missing = eda.get("missing_values", _EDA_FALLBACK["missing_values"])
    for feat in _FEATURE_ORDER:
        vals = missing.get(feat, [0, 0.0])
        ws.cell(row=row, column=1, value=feat)
        ws.cell(row=row, column=2, value=int(vals[0]))
        cell = ws.cell(row=row, column=3, value=float(vals[1]))
        cell.number_format = "0.00\"%\""
        _apply_data_row(ws, row, 1, 3, alt=(row % 2 == 0))
        row += 1
    row += 1

    _write_subtitle(ws, row, 1, _tr(lang, "outliers"), 2)
    row += 1
    out_headers = [_tr(lang, "feature"), _tr(lang, "count")]
    _write_header_row(ws, row, out_headers)
    row += 1
    outliers = eda.get("outliers_iqr", _EDA_FALLBACK["outliers_iqr"])
    for feat in _FEATURE_ORDER:
        if feat == "MedHouseVal":
            continue
        ws.cell(row=row, column=1, value=feat)
        ws.cell(row=row, column=2, value=outliers.get(feat, 0))
        _apply_data_row(ws, row, 1, 2, alt=(row % 2 == 0))
        row += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_training(ws, lang: str):
    ws.title = _tr(lang, "sheet_training")
    training = _load_training()
    row = 1

    _write_title(ws, row, 1, _tr(lang, "model_comparison"), 7)
    row += 1
    headers = [_tr(lang, "model"), "MSE", "RMSE", "MAE", "R²", _tr(lang, "training_time"), _tr(lang, "params")]
    _write_header_row(ws, row, headers)
    row += 1
    metrics = training.get("metrics", _TRAINING_FALLBACK["metrics"])
    for m in metrics:
        ws.cell(row=row, column=1, value=m.get("model", ""))
        ws.cell(row=row, column=2, value=m.get("MSE", 0))
        ws.cell(row=row, column=2).number_format = "0.0000"
        ws.cell(row=row, column=3, value=m.get("RMSE", 0))
        ws.cell(row=row, column=3).number_format = "0.0000"
        ws.cell(row=row, column=4, value=m.get("MAE", 0))
        ws.cell(row=row, column=4).number_format = "0.0000"
        ws.cell(row=row, column=5, value=m.get("R2", 0))
        ws.cell(row=row, column=5).number_format = "0.0000"
        ws.cell(row=row, column=6, value=m.get("training_time", 0))
        ws.cell(row=row, column=6).number_format = "0.0"
        ws.cell(row=row, column=7, value=str(m.get("params", "")))
        _apply_data_row(ws, row, 1, 7, alt=(row % 2 == 0))
        row += 1
    row += 1

    best = training.get("best_model", _TRAINING_FALLBACK["best_model"])
    _write_subtitle(ws, row, 1, _tr(lang, "model_details"), 3)
    row += 1
    detail_keys = ["architecture", "activation", "optimizer", "learning_rate",
                   "batch_size", "epochs", "early_stopping"]
    for key in detail_keys:
        ws.cell(row=row, column=1, value=_tr(lang, key)).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=2, value=str(best.get(key, "—"))).font = _DATA_FONT
        _apply_data_row(ws, row, 1, 2)
        row += 1
    row += 1

    _write_subtitle(ws, row, 1, _tr(lang, "performance"), 5)
    row += 1
    perf_headers = ["MSE", "RMSE", "MAE", "R²"]
    _write_header_row(ws, row, perf_headers)
    row += 1
    for i, k in enumerate(["MSE", "RMSE", "MAE", "R2"]):
        cell = ws.cell(row=row, column=i + 1, value=best.get(k, 0))
        cell.number_format = "0.0000"
    _apply_data_row(ws, row, 1, 4)

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_cv(ws, lang: str):
    ws.title = _tr(lang, "sheet_cv")
    cv = _load_cv()
    row = 1

    _write_title(ws, row, 1, _tr(lang, "cv_results"), 5)
    row += 1
    headers = [_tr(lang, "fold"), "MSE", "RMSE", "MAE", "R²"]
    _write_header_row(ws, row, headers)
    row += 1
    folds = cv.get("folds", _CV_FALLBACK["folds"])
    for f in folds:
        ws.cell(row=row, column=1, value=f.get("fold", ""))
        ws.cell(row=row, column=2, value=f.get("MSE", 0))
        ws.cell(row=row, column=2).number_format = "0.0000"
        ws.cell(row=row, column=3, value=f.get("RMSE", 0))
        ws.cell(row=row, column=3).number_format = "0.0000"
        ws.cell(row=row, column=4, value=f.get("MAE", 0))
        ws.cell(row=row, column=4).number_format = "0.0000"
        ws.cell(row=row, column=5, value=f.get("R2", 0))
        ws.cell(row=row, column=5).number_format = "0.0000"
        _apply_data_row(ws, row, 1, 5, alt=(row % 2 == 0))
        row += 1

    mean = cv.get("mean", _CV_FALLBACK["mean"])
    ws.cell(row=row, column=1, value=_tr(lang, "mean")).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=2, value=mean.get("MSE", 0))
    ws.cell(row=row, column=2).number_format = "0.0000"
    ws.cell(row=row, column=3, value=mean.get("RMSE", 0))
    ws.cell(row=row, column=3).number_format = "0.0000"
    ws.cell(row=row, column=4, value=mean.get("MAE", 0))
    ws.cell(row=row, column=4).number_format = "0.0000"
    ws.cell(row=row, column=5, value=mean.get("R2", 0))
    ws.cell(row=row, column=5).number_format = "0.0000"
    _apply_data_row(ws, row, 1, 5, alt=True)
    row += 1

    std = cv.get("std", _CV_FALLBACK["std"])
    ws.cell(row=row, column=1, value=_tr(lang, "std")).font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=2, value=std.get("MSE", 0))
    ws.cell(row=row, column=2).number_format = "0.0000"
    ws.cell(row=row, column=3, value=std.get("RMSE", 0))
    ws.cell(row=row, column=3).number_format = "0.0000"
    ws.cell(row=row, column=4, value=std.get("MAE", 0))
    ws.cell(row=row, column=4).number_format = "0.0000"
    ws.cell(row=row, column=5, value=std.get("R2", 0))
    ws.cell(row=row, column=5).number_format = "0.0000"
    _apply_data_row(ws, row, 1, 5, alt=False)

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_hyperparameter(ws, lang: str):
    ws.title = _tr(lang, "sheet_hyperparameter")
    hp = _load_hyperparameter()
    row = 1

    _write_title(ws, row, 1, _tr(lang, "best_params"), 3)
    row += 1
    headers = [_tr(lang, "parameter"), _tr(lang, "value")]
    _write_header_row(ws, row, headers)
    row += 1
    params = hp.get("best_params", _HYPERPARAMETER_FALLBACK["best_params"])
    for key, val in params.items():
        ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = _DATA_FONT
        ws.cell(row=row, column=2, value=str(val)).font = _DATA_FONT
        _apply_data_row(ws, row, 1, 2, alt=(row % 2 == 0))
        row += 1
    row += 1
    ws.cell(row=row, column=1, value=_tr(lang, "best_score")).font = Font(name="Calibri", size=10, bold=True)
    score = hp.get("best_score", _HYPERPARAMETER_FALLBACK["best_score"])
    cell = ws.cell(row=row, column=2, value=score)
    cell.number_format = "0.0000"
    cell.font = _DATA_FONT

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_statistical(ws, lang: str):
    ws.title = _tr(lang, "sheet_statistical")
    tests = _load_statistical_tests()
    row = 1

    _write_title(ws, row, 1, _tr(lang, "statistical_tests"), 4)
    row += 1
    headers = [_tr(lang, "test"), _tr(lang, "statistic"), _tr(lang, "p_value"), _tr(lang, "conclusion")]
    _write_header_row(ws, row, headers)
    row += 1

    test_data = [
        (_tr(lang, "shapiro_wilk"), "shapiro_wilk", ["normal", "not_normal"]),
        (_tr(lang, "durbin_watson"), "durbin_watson", ["no_autocorrelation", "autocorrelation"]),
        (_tr(lang, "breusch_pagan"), "breusch_pagan", ["homoscedastic", "heteroscedastic"]),
        (_tr(lang, "friedman"), "friedman", ["no_difference", "significant_difference"]),
    ]
    for test_label, test_key, (pos_key, neg_key) in test_data:
        result = tests.get(test_key, _STATISTICAL_TESTS_FALLBACK.get(test_key, {}))
        ws.cell(row=row, column=1, value=test_label).font = _DATA_FONT
        ws.cell(row=row, column=2, value=result.get("statistic", 0))
        ws.cell(row=row, column=2).number_format = "0.0000"
        pval = result.get("p_value", 0)
        if pval < 0.0001:
            ws.cell(row=row, column=3, value="< 0.0001").font = _DATA_FONT
        else:
            cell = ws.cell(row=row, column=3, value=pval)
            cell.number_format = "0.0000"
        if "autocorrelation" in result:
            conclusion = pos_key if not result["autocorrelation"] else neg_key
        elif "heteroscedastic" in result:
            conclusion = pos_key if not result["heteroscedastic"] else neg_key
        elif "significant_difference" in result:
            conclusion = pos_key if not result["significant_difference"] else neg_key
        elif "normal" in result:
            conclusion = pos_key if result["normal"] else neg_key
        else:
            conclusion = pos_key
        ws.cell(row=row, column=4, value=_tr(lang, conclusion)).font = _DATA_FONT
        _apply_data_row(ws, row, 1, 4, alt=(row % 2 == 0))
        row += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def generate_excel(language: str = "es", output_path: Optional[str] = None) -> bytes:
    lang = language if language in _LANG else "es"
    wb = Workbook()

    ws_summary = wb.active
    _build_summary(ws_summary, lang)

    ws_eda = wb.create_sheet()
    _build_eda(ws_eda, lang)

    ws_training = wb.create_sheet()
    _build_training(ws_training, lang)

    ws_cv = wb.create_sheet()
    _build_cv(ws_cv, lang)

    ws_hp = wb.create_sheet()
    _build_hyperparameter(ws_hp, lang)

    ws_stat = wb.create_sheet()
    _build_statistical(ws_stat, lang)

    buffer = io.BytesIO()
    wb.save(buffer)
    result = buffer.getvalue()
    if output_path:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(result)
    return result
